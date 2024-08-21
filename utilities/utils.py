import logging
import inspect
import softest
import openpyxl 

class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            print('The text is' + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("Test Passed")
            else:
                print("Test Failed")
        self.assert_all()

    def custom_Logger(logLevel = logging.DEBUG):
        # set class/method name from where it's called
        logger_name = inspect.stack()[1][3]

        logger = logging.getLogger((logger_name))
        logger.setLevel(logLevel)
        # create console handler/file handler
        fh = logging.FileHandler("automation.log", mode='a')
        #create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')

        # add formmatter to console handler
        fh.setFormatter(formatter)

        # add console handler to logger
        logger.addHandler(fh)
        return logger
    
    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet]
        row_ct = sheet.max_row
        col_ct = sheet.max_column
        
        for i in range(1, row_ct+1):
            row =[]
            for j in range(1, col_ct+1):
                row.append(sheet.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist